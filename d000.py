import sys
import re
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Color

folder = Path(sys.argv[1]).resolve()
regex = re.compile(r"\\\d{3}")
wb = openpyxl.load_workbook("c:/users/crazy/dropbox/share/이상한 환상향 크로니클 번역.xlsx")
ws = wb["d000"]

def convert(string):
    byteList = bytes(0)
    i=0
    while(i<len(string)):
        if regex.match(string[i:i+4]):
            byteList += int(string[i+1:i+4]).to_bytes(1,"big")
            i+=4
        else:
            byteList += ord(string[i]).to_bytes(1,"big")
            i+=1
    return byteList.decode("cp932")

c=2
(newPath := folder.with_name("d000_")).mkdir()
for file in [i for i in folder.rglob("*") if i.is_file()]:
    data = file.read_bytes()
    if not regex.search(data.decode("utf-8")):
        continue

    print(file)
    ws.cell(c,4).value = str(file.name)
    ws.cell(c,4).fill = PatternFill(patternType="solid",fgColor=Color("7AD694"))
    c+=1
    result = re.compile("\".+?\"").findall(data.decode("utf-8"))
    result = [convert(i[1:-1]) for i in result if regex.search(i)]
    for r in result:
        ws.cell(c,4).value = r
        c+=1
        """
        if not os.path.exists(newPath):
            os.makedirs(newPath)
        f = open(newPath+"/"+file,"wb")
        i=0
        while(i<len(data)):
            result = regex.match(data[i:i+4].decode("utf-8"))
            if result:
                f.write(bytes([int(result.group()[1:])]))
                i+=4
            else:
                f.write(data[i:i+1])
                i+=1
        f.close()
        """
wb.save("c:/users/crazy/dropbox/share/이상한 환상향 크로니클 번역.xlsx")
wb.close()